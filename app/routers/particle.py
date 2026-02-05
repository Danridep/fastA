from fastapi import APIRouter, HTTPException, UploadFile, File, Query
from typing import Optional
import openpyxl
import os
from datetime import datetime
from app.database import get_db_cursor
from app.models import BaseResponse, PaginatedResponse
from app.utils.file_utils import save_uploaded_file, delete_file
from app.config import settings

router = APIRouter()


def get_cell_ref(row, col):
    """Получить ссылку на ячейку в формате R1C1"""
    return f"R{row}C{col}"


@router.post("/compare")
async def compare_excel_files(
        file1: UploadFile = File(...),
        file2: UploadFile = File(...)
):
    """
    Сравнить два Excel файла и найти минусовые значения
    """
    try:
        # Проверка расширений файлов
        if not (file1.filename.lower().endswith(('.xlsx', '.xls')) and
                file2.filename.lower().endswith(('.xlsx', '.xls'))):
            raise HTTPException(status_code=400, detail="Оба файла должны быть в формате Excel")

        # Создаем директории если их нет
        os.makedirs(settings.UPLOAD_DIR, exist_ok=True)

        # Сохраняем загруженные файлы
        file1_path = save_uploaded_file(file1, settings.UPLOAD_DIR)
        file2_path = save_uploaded_file(file2, settings.UPLOAD_DIR)

        try:
            # Загружаем файлы
            wb1 = openpyxl.load_workbook(file1_path, data_only=True)
            ws1 = wb1.active

            wb2 = openpyxl.load_workbook(file2_path, data_only=True)
            ws2 = wb2.active

            # --- Поиск минусов в файле 1 (по всем ячейкам) ---
            minus_file1 = []
            for row in ws1.iter_rows(min_row=2, values_only=False):  # пропускаем заголовки
                for cell in row:
                    if isinstance(cell.value, (int, float)) and cell.value < 0:
                        minus_file1.append({
                            "cell": get_cell_ref(cell.row, cell.column),
                            "value": cell.value,
                            "row": cell.row,
                            "column": cell.column
                        })

            # --- Поиск минусов в файле 2 (только столбцы 3 и 5) ---
            minus_file2 = []
            for row in ws2.iter_rows(min_row=2, values_only=False):
                # Столбец 3 (индекс 2)
                cell_col3 = row[2] if len(row) > 2 else None
                if cell_col3 and isinstance(cell_col3.value, (int, float)) and cell_col3.value < 0:
                    minus_file2.append({
                        "cell": get_cell_ref(cell_col3.row, 3),
                        "value": cell_col3.value,
                        "row": cell_col3.row,
                        "column": 3
                    })

                # Столбец 5 (индекс 4)
                cell_col5 = row[4] if len(row) > 4 else None
                if cell_col5 and isinstance(cell_col5.value, (int, float)) and cell_col5.value < 0:
                    minus_file2.append({
                        "cell": get_cell_ref(cell_col5.row, 5),
                        "value": cell_col5.value,
                        "row": cell_col5.row,
                        "column": 5
                    })

            # --- Итоговое значение в файле 2 ---
            total2 = None
            total2_cell = None
            for row in ws2.iter_rows(min_row=1, values_only=False):
                if len(row) > 1 and row[1].value and str(row[1].value).strip().lower() == "итог":
                    total2 = row[5].value if len(row) > 5 else None  # 6-й столбец (индекс 5)
                    total2_cell = get_cell_ref(row[1].row, 6) if total2 is not None else None
                    break

            # --- Итоговое значение в файле 1 ---
            total1 = None
            total1_cell = None
            for row in ws1.iter_rows(min_row=1, values_only=False):
                if len(row) > 1 and row[1].value and str(row[1].value).strip().lower() == "итог":
                    total1 = row[8].value if len(row) > 8 else None  # 9-й столбец (индекс 8)
                    total1_cell = get_cell_ref(row[1].row, 9) if total1 is not None else None
                    break

            # --- Сравнение итогов ---
            comparison = "match" if total1 == total2 else "mismatch"
            comparison_message = "Сошлись" if comparison == "match" else "Не сошлись"

            # Сохраняем в историю
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            with get_db_cursor() as cursor:
                cursor.execute(
                    """
                    INSERT INTO particle_history 
                    (file1_name, file2_name, total1, total2, comparison, 
                     minus_count1, minus_count2, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (file1.filename, file2.filename, total1, total2, comparison,
                     len(minus_file1), len(minus_file2), timestamp)
                )
                history_id = cursor.lastrowid

            result = {
                "history_id": history_id,
                "file1": {
                    "name": file1.filename,
                    "minus_cells": minus_file1,
                    "minus_count": len(minus_file1),
                    "total": total1,
                    "total_cell": total1_cell
                },
                "file2": {
                    "name": file2.filename,
                    "minus_cells": minus_file2,
                    "minus_count": len(minus_file2),
                    "total": total2,
                    "total_cell": total2_cell
                },
                "comparison": {
                    "status": comparison,
                    "message": comparison_message,
                    "total1": total1,
                    "total2": total2
                },
                "timestamp": timestamp
            }

            return BaseResponse(data=result)

        finally:
            # Удаляем временные файлы
            for file_path in [file1_path, file2_path]:
                if os.path.exists(file_path):
                    delete_file(file_path)

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка при сравнении файлов: {str(e)}")


@router.get("/history", response_model=PaginatedResponse)
async def get_comparison_history(
        page: int = Query(1, ge=1, description="Номер страницы"),
        per_page: int = Query(20, ge=1, le=100, description="Количество записей на странице"),
        search: Optional[str] = Query(None, description="Поиск по именам файлов"),
        comparison: Optional[str] = Query(None, description="Фильтр по результату (match/mismatch)"),
        date_from: Optional[str] = Query(None, description="Дата от (формат: ГГГГ-ММ-ДД)"),
        date_to: Optional[str] = Query(None, description="Дата до (формат: ГГГГ-ММ-ДД)")
):
    """
    Получить историю сравнений с пагинацией и фильтрами
    """
    try:
        print(f"DEBUG: Запрос истории сравнений. Страница: {page}, Размер: {per_page}")

        with get_db_cursor() as cursor:
            # Сначала проверим, есть ли таблица
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='particle_history'")
            table_exists = cursor.fetchone()

            if not table_exists:
                print("DEBUG: Таблица particle_history не существует!")
                return PaginatedResponse(
                    success=True,
                    message="Таблица истории не существует",
                    total=0,
                    page=page,
                    pages=0,
                    per_page=per_page,
                    data=[]
                )

            # Проверим, есть ли данные в таблице
            cursor.execute("SELECT COUNT(*) as cnt FROM particle_history")
            count_result = cursor.fetchone()
            print(f"DEBUG: Всего записей в таблице: {count_result}")

            # Формируем базовый запрос и условия
            query = "SELECT * FROM particle_history WHERE 1=1"
            count_query = "SELECT COUNT(*) as total FROM particle_history WHERE 1=1"
            params = []

            # Добавляем условия поиска
            if search:
                query += " AND (file1_name LIKE ? OR file2_name LIKE ?)"
                count_query += " AND (file1_name LIKE ? OR file2_name LIKE ?)"
                search_param = f"%{search}%"
                params.extend([search_param, search_param])

            # Фильтр по результату сравнения
            if comparison and comparison in ["match", "mismatch"]:
                query += " AND comparison = ?"
                count_query += " AND comparison = ?"
                params.append(comparison)

            # Фильтр по дате
            if date_from:
                query += " AND DATE(created_at) >= ?"
                count_query += " AND DATE(created_at) >= ?"
                params.append(date_from)

            if date_to:
                query += " AND DATE(created_at) <= ?"
                count_query += " AND DATE(created_at) <= ?"
                params.append(date_to)

            # Получаем общее количество записей
            print(f"DEBUG: Выполняем count_query: {count_query} с params: {params}")
            cursor.execute(count_query, params)
            total_result = cursor.fetchone()
            total = total_result["total"] if total_result and "total" in total_result else 0
            print(f"DEBUG: Найдено записей: {total}")

            # Добавляем сортировку и пагинацию
            query += " ORDER BY created_at DESC LIMIT ? OFFSET ?"

            # Рассчитываем смещение
            offset = (page - 1) * per_page
            query_params = params.copy()
            query_params.extend([per_page, offset])

            # Выполняем запрос с пагинацией
            print(f"DEBUG: Выполняем query: {query} с params: {query_params}")
            cursor.execute(query, query_params)
            items = cursor.fetchall()
            print(f"DEBUG: Получено записей: {len(items)}")

            # Форматируем даты для удобства отображения
            formatted_items = []
            for item in items:
                formatted_item = dict(item)  # Конвертируем строку в словарь

                try:
                    # Пробуем разные форматы даты
                    created_at_str = formatted_item["created_at"]
                    if created_at_str:
                        # Пробуем парсить как datetime
                        try:
                            created_at = datetime.strptime(created_at_str, "%Y-%m-%d %H:%M:%S")
                        except:
                            try:
                                created_at = datetime.strptime(created_at_str, "%Y-%m-%d")
                            except:
                                created_at = datetime.now()

                        formatted_item["created_at_formatted"] = created_at.strftime("%d.%m.%Y %H:%M")
                        formatted_item["date_only"] = created_at.strftime("%d.%m.%Y")
                        formatted_item["time_only"] = created_at.strftime("%H:%M")
                    else:
                        formatted_item["created_at_formatted"] = "Нет данных"
                        formatted_item["date_only"] = ""
                        formatted_item["time_only"] = ""
                except Exception as e:
                    print(f"DEBUG: Ошибка форматирования даты: {e}")
                    formatted_item["created_at_formatted"] = formatted_item.get("created_at", "Нет данных")
                    formatted_item["date_only"] = ""
                    formatted_item["time_only"] = ""

                # Добавляем русское название результата
                comparison_value = formatted_item.get("comparison", "")
                formatted_item["comparison_ru"] = "Сошлись" if comparison_value == "match" else "Не сошлись"
                formatted_item["comparison_class"] = "success" if comparison_value == "match" else "danger"
                formatted_item["comparison_icon"] = "fa-check" if comparison_value == "match" else "fa-times"

                formatted_items.append(formatted_item)

            # Вычисляем количество страниц
            pages = (total + per_page - 1) // per_page if per_page > 0 else 1

            return PaginatedResponse(
                success=True,
                message=f"Загружено {len(formatted_items)} из {total} записей",
                total=total,
                page=page,
                pages=pages,
                per_page=per_page,
                data=formatted_items
            )

    except Exception as e:
        import traceback
        print(f"ERROR: Ошибка загрузки истории: {str(e)}")
        print(f"DEBUG: {traceback.format_exc()}")
        return PaginatedResponse(
            success=False,
            message=f"Ошибка загрузки истории: {str(e)}",
            total=0,
            page=page,
            pages=0,
            per_page=per_page,
            data=[]
        )


@router.get("/history/summary", response_model=BaseResponse)
async def get_history_summary():
    """
    Получить сводную информацию по истории сравнений
    """
    try:
        with get_db_cursor() as cursor:
            # Проверяем существование таблицы
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='particle_history'")
            if not cursor.fetchone():
                return BaseResponse(data={
                    "summary": {"total": 0, "matches": 0, "mismatches": 0},
                    "last_30_days": [],
                    "top_files1": [],
                    "top_files2": [],
                    "recent": []
                })

            # Общая статистика
            cursor.execute("""
                SELECT 
                    COUNT(*) as total,
                    SUM(CASE WHEN comparison = 'match' THEN 1 ELSE 0 END) as matches,
                    SUM(CASE WHEN comparison = 'mismatch' THEN 1 ELSE 0 END) as mismatches
                FROM particle_history
            """)
            summary = cursor.fetchone() or {"total": 0, "matches": 0, "mismatches": 0}

            # Статистика по последним 30 дням
            cursor.execute("""
                SELECT 
                    DATE(created_at) as date,
                    COUNT(*) as count
                FROM particle_history 
                WHERE created_at >= DATE('now', '-30 days')
                GROUP BY DATE(created_at)
                ORDER BY date DESC
            """)
            last_30_days = cursor.fetchall() or []

            # Самые частые файлы
            cursor.execute("""
                SELECT 
                    file1_name,
                    COUNT(*) as count
                FROM particle_history 
                GROUP BY file1_name
                ORDER BY count DESC
                LIMIT 5
            """)
            top_files1 = cursor.fetchall() or []

            cursor.execute("""
                SELECT 
                    file2_name,
                    COUNT(*) as count
                FROM particle_history 
                GROUP BY file2_name
                ORDER BY count DESC
                LIMIT 5
            """)
            top_files2 = cursor.fetchall() or []

            # Последние 5 записей
            cursor.execute("""
                SELECT * FROM particle_history 
                ORDER BY created_at DESC 
                LIMIT 5
            """)
            recent = cursor.fetchall() or []

            return BaseResponse(data={
                "summary": dict(summary),
                "last_30_days": last_30_days,
                "top_files1": top_files1,
                "top_files2": top_files2,
                "recent": recent
            })

    except Exception as e:
        print(f"ERROR: Ошибка загрузки сводной информации: {e}")
        return BaseResponse(
            success=False,
            message=f"Ошибка загрузки сводной информации: {str(e)}",
            data={}
        )


@router.get("/history/{history_id}", response_model=BaseResponse)
async def get_comparison_details(history_id: int):
    """
    Получить детали сравнения по ID
    """
    try:
        with get_db_cursor() as cursor:
            cursor.execute(
                "SELECT * FROM particle_history WHERE id = ?",
                (history_id,)
            )
            history_item = cursor.fetchone()

            if not history_item:
                raise HTTPException(status_code=404, detail="Запись истории не найдена")

            # Конвертируем в словарь
            history_dict = dict(history_item)

            # Форматируем дату
            try:
                created_at_str = history_dict["created_at"]
                if created_at_str:
                    try:
                        created_at = datetime.strptime(created_at_str, "%Y-%m-%d %H:%M:%S")
                    except:
                        created_at = datetime.strptime(created_at_str, "%Y-%m-%d")

                    history_dict["created_at_formatted"] = created_at.strftime("%d.%m.%Y %H:%M:%S")
                    history_dict["date_only"] = created_at.strftime("%d.%m.%Y")
                    history_dict["time_only"] = created_at.strftime("%H:%M:%S")
                else:
                    history_dict["created_at_formatted"] = "Нет данных"
            except:
                history_dict["created_at_formatted"] = history_dict.get("created_at", "Нет данных")

            # Добавляем русское название результата
            comparison_value = history_dict.get("comparison", "")
            history_dict["comparison_ru"] = "Сошлись" if comparison_value == "match" else "Не сошлись"
            history_dict["comparison_class"] = "success" if comparison_value == "match" else "danger"

            return BaseResponse(data=history_dict)

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка загрузки деталей: {str(e)}")


@router.delete("/history/{history_id}", response_model=BaseResponse)
async def delete_comparison_history(history_id: int):
    """
    Удалить запись из истории сравнений
    """
    try:
        with get_db_cursor() as cursor:
            cursor.execute(
                "DELETE FROM particle_history WHERE id = ?",
                (history_id,)
            )

            if cursor.rowcount == 0:
                raise HTTPException(status_code=404, detail="Запись истории не найдена")

            return BaseResponse(
                success=True,
                message="Запись истории удалена успешно"
            )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка удаления истории: {str(e)}")


@router.get("/stats", response_model=BaseResponse)
async def get_comparison_stats():
    """
    Получить статистику сравнений
    """
    try:
        with get_db_cursor() as cursor:
            # Проверяем существование таблицы
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='particle_history'")
            if not cursor.fetchone():
                return BaseResponse(data={
                    "overall": {
                        "total_comparisons": 0,
                        "matches": 0,
                        "mismatches": 0,
                        "first_comparison": None,
                        "last_comparison": None
                    },
                    "last_7_days": [],
                    "success_rate": 0
                })

            cursor.execute("""
                SELECT 
                    COUNT(*) as total_comparisons,
                    SUM(CASE WHEN comparison = 'match' THEN 1 ELSE 0 END) as matches,
                    SUM(CASE WHEN comparison = 'mismatch' THEN 1 ELSE 0 END) as mismatches,
                    MIN(created_at) as first_comparison,
                    MAX(created_at) as last_comparison,
                    AVG(minus_count1) as avg_minus1,
                    AVG(minus_count2) as avg_minus2,
                    MAX(minus_count1) as max_minus1,
                    MAX(minus_count2) as max_minus2
                FROM particle_history
            """)
            stats = cursor.fetchone() or {}

            # Статистика по дням
            cursor.execute("""
                SELECT 
                    DATE(created_at) as date,
                    COUNT(*) as count,
                    SUM(CASE WHEN comparison = 'match' THEN 1 ELSE 0 END) as matches,
                    SUM(CASE WHEN comparison = 'mismatch' THEN 1 ELSE 0 END) as mismatches
                FROM particle_history 
                GROUP BY DATE(created_at)
                ORDER BY date DESC
                LIMIT 7
            """)
            last_7_days = cursor.fetchall() or []

            # Рассчитываем процент успеха
            total = stats.get("total_comparisons", 0)
            matches = stats.get("matches", 0)
            success_rate = round((matches / total * 100) if total > 0 else 0, 2)

            return BaseResponse(data={
                "overall": dict(stats),
                "last_7_days": last_7_days,
                "success_rate": success_rate
            })

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка загрузки статистики: {str(e)}")