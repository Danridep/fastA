"""
Сервис генерации текста из Excel-файла по настраиваемому паттерну.
"""
import re
import io
from typing import Any

try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    openpyxl = None


def _get_cell_value(cell) -> str:
    """Получить строковое значение ячейки."""
    if cell is None or cell.value is None:
        return ""
    v = cell.value
    if isinstance(v, float):
        # Убираем лишний .0 у целых
        return str(int(v)) if v == int(v) else str(v)
    return str(v).strip()


def parse_excel_preview(file_bytes: bytes, sheet_index: int = 0,
                        header_row: int = 1, preview_rows: int = 5) -> dict:
    """
    Парсит Excel и возвращает:
    - columns: список {index, name} (из header_row)
    - rows: первые preview_rows строк данных
    - total_rows: общее кол-во строк данных
    - sheets: список имён листов
    """
    if openpyxl is None:
        raise RuntimeError("openpyxl не установлен. Выполни: pip install openpyxl")

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheets = wb.sheetnames

    if sheet_index >= len(sheets):
        sheet_index = 0

    ws = wb.worksheets[sheet_index]
    rows = list(ws.iter_rows())

    if not rows:
        return {"columns": [], "rows": [], "total_rows": 0, "sheets": sheets}

    # Заголовки
    header_idx = header_row - 1
    if header_idx >= len(rows):
        header_idx = 0

    header_cells = rows[header_idx]
    columns = []
    for i, cell in enumerate(header_cells):
        v = _get_cell_value(cell)
        columns.append({
            "index": i,            # 0-based
            "col_letter": cell.column_letter if hasattr(cell, 'column_letter') else str(i+1),
            "name": v or f"Столбец {i+1}",
            "original": v
        })

    # Данные
    data_rows = rows[header_idx + 1:]
    total_rows = len(data_rows)
    max_col = len(columns)

    preview = []
    for row in data_rows[:preview_rows]:
        cells = [_get_cell_value(c) for c in row[:max_col]]
        # Дополняем если строка короче заголовка
        while len(cells) < max_col:
            cells.append("")
        preview.append(cells)

    wb.close()
    return {
        "columns": columns,
        "rows": preview,
        "total_rows": total_rows,
        "sheets": sheets
    }


def generate_text(file_bytes: bytes, config: dict) -> dict:
    """
    Генерирует текст по конфигу.

    config:
        sheet_index    int   (0-based, default 0)
        header_row     int   (1-based, default 1)
        skip_rows      int   (сколько строк пропустить после заголовка)
        pattern        str   шаблон вида "{0} {Адрес} - {ФИО}"
        separator      str   разделитель между записями (default "\n")
        skip_empty     bool  пропускать строки где все выбранные поля пусты
        filters        list  [{col, value, mode}] mode: contains/equals/not_empty
        max_rows       int   ограничение (0 = без ограничения)
    """
    if openpyxl is None:
        raise RuntimeError("openpyxl не установлен")

    sheet_index = int(config.get("sheet_index", 0))
    header_row  = int(config.get("header_row", 1))
    skip_rows   = int(config.get("skip_rows", 0))
    pattern     = config.get("pattern", "")
    separator   = config.get("separator", "\n")
    skip_empty  = bool(config.get("skip_empty", True))
    filters     = config.get("filters", [])
    max_rows    = int(config.get("max_rows", 0))

    # Заменяем \n в сепараторе
    separator = separator.replace("\\n", "\n")

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.worksheets[min(sheet_index, len(wb.worksheets) - 1)]
    rows = list(ws.iter_rows())
    wb.close()

    if not rows:
        return {"text": "", "count": 0, "skipped": 0}

    header_idx   = header_row - 1
    header_cells = rows[header_idx] if header_idx < len(rows) else rows[0]

    # Строим маппинг: имя столбца → индекс
    col_name_to_idx = {}
    col_names = []
    for i, cell in enumerate(header_cells):
        v = _get_cell_value(cell) or f"Столбец {i+1}"
        col_name_to_idx[v] = i
        col_name_to_idx[str(i)] = i          # и по числовому индексу
        col_names.append(v)

    # Находим все плейсхолдеры в паттерне: {ИмяСтолбца} или {0}
    placeholders = re.findall(r'\{([^}]+)\}', pattern)

    data_rows = rows[header_idx + 1 + skip_rows:]

    results = []
    skipped = 0
    max_col = len(col_names)

    for raw_row in data_rows:
        if max_rows and len(results) >= max_rows:
            break

        cells = [_get_cell_value(c) for c in raw_row[:max_col]]
        while len(cells) < max_col:
            cells.append("")

        # Применяем фильтры
        if filters:
            skip = False
            for f in filters:
                col_ref = f.get("col", "")
                idx = col_name_to_idx.get(col_ref, col_name_to_idx.get(str(col_ref)))
                if idx is None:
                    continue
                val = cells[idx] if idx < len(cells) else ""
                mode = f.get("mode", "not_empty")
                fval = f.get("value", "")
                if mode == "not_empty" and not val:
                    skip = True; break
                if mode == "equals" and val.lower() != fval.lower():
                    skip = True; break
                if mode == "contains" and fval.lower() not in val.lower():
                    skip = True; break
                if mode == "not_contains" and fval.lower() in val.lower():
                    skip = True; break
            if skip:
                skipped += 1
                continue

        # Строим замены
        values_used = []
        row_values = {}
        for ph in placeholders:
            idx = col_name_to_idx.get(ph)
            if idx is None:
                # Пробуем как число
                try:
                    idx = int(ph)
                except ValueError:
                    row_values[ph] = f"[?{ph}]"
                    continue
            val = cells[idx] if idx < len(cells) else ""
            row_values[ph] = val
            values_used.append(val)

        # skip_empty: пропускаем если все используемые поля пусты
        if skip_empty and all(v == "" for v in values_used):
            skipped += 1
            continue

        # Применяем паттерн
        line = pattern
        for ph, val in row_values.items():
            line = line.replace(f"{{{ph}}}", val)

        results.append(line)

    text = separator.join(results)
    return {
        "text": text,
        "count": len(results),
        "skipped": skipped,
        "total_processed": len(results) + skipped
    }