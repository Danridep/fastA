import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Dict, Any, List
import io


def create_order_excel(data: Dict[str, Any], order_type: str) -> bytes:
    """
    Создать Excel файл для заказа
    """
    # Создаем новую рабочую книгу
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заполненный заказ"

    headers = data["headers"]
    addresses_data = data["addresses_data"]

    # Стили
    header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    even_row_fill = PatternFill(start_color="F7F9FC", end_color="F7F9FC", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Записываем заголовки
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    # Записываем данные
    row_idx = 2
    for address in data["addresses"]:
        if address in addresses_data:
            for row_data in addresses_data[address]:
                # Пропускаем строки с пустым количеством
                if not row_data.get("Кол-во"):
                    continue

                for col_idx, header in enumerate(headers, 1):
                    value = row_data.get(header, "")
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border

                    # Чередование строк
                    if row_idx % 2 == 0:
                        cell.fill = even_row_fill

                    # Выравнивание для числовых полей
                    if header == "Кол-во":
                        cell.alignment = Alignment(horizontal="center")
                        try:
                            if str(value).isdigit():
                                cell.number_format = '0'
                        except:
                            pass

                row_idx += 1

    # Автоширина колонок
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width

    # Добавляем итоговую строку
    if row_idx > 2:
        total_row = row_idx + 1
        ws.cell(row=total_row, column=1, value="ИТОГО:").font = Font(bold=True)

        # Сумма количества
        quantity_col = None
        for idx, header in enumerate(headers, 1):
            if header == "Кол-во":
                quantity_col = idx
                break

        if quantity_col:
            formula = f"=SUM({get_column_letter(quantity_col)}2:{get_column_letter(quantity_col)}{row_idx})"
            ws.cell(row=total_row, column=quantity_col, value=formula).font = Font(bold=True)

    # Сохраняем в bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()