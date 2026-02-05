import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from collections import defaultdict
import warnings
import re
import os
from typing import List, Dict, Any, Optional
from datetime import datetime
import json
from fastapi import HTTPException
from app.database import get_db_cursor

warnings.filterwarnings('ignore')


class WorkOrderProcessor:
    def __init__(self, use_database: bool = True):
        self.work_types_df = None
        self.orders_df = None
        self.use_database = use_database

        # Заявки с нарушениями
        self.violation_orders = []
        self.to_violations = []  # ТО заявки с нарушениями
        self.non_to_violations = []  # Не-ТО заявки с нарушениями

        # Словарь типов работ из базы
        self.work_types_dict = {}

        # Украинские области и города
        self.regions = {
            'Кременчук і Горішні Плавні (окремо)': [
                'Кременчук', 'Горішні Плавні'
            ],
            'Полтавська область': [
                'Полтава', 'Лубни', 'Миргород', 'Розсошенці'
            ],
            'Одеська область': [
                'Одеса', 'Подільськ', 'Теплодар', 'Чорноморськ',
                'Южне', 'Білгород-Дністровський', 'Ізмаїл',
                'Сергіївка', 'Арциз'
            ],
            'Миколаївська область': [
                'Миколаїв', 'Первомайськ', 'Вознесенськ', 'Південноукраїнськ'
            ],
            'Черкаська область': [
                'Черкаси', 'Сміла', 'Монастирище', 'Умань'
            ],
            'Кіровоградська область': [
                'Кропивницький', 'Світловодськ', 'Олександрія',
                'Долинська', 'Смоліне'
            ],
            'Київська область': [
                'Київ', 'Бровари', 'Біла Церква', 'Ірпінь', 'Бориспіль'
            ],
            'Харківська область': [
                'Харків', 'Чугуїв', 'Ізюм', 'Куп\'янськ'
            ],
            'Дніпропетровська область': [
                'Дніпро', 'Кривий Ріг', 'Нікополь', 'Павлоград', 'Каменське'
            ],
            'Львівська область': [
                'Львів', 'Дрогобич', 'Червоноград', 'Стрий'
            ],
            'Запорізька область': [
                'Запоріжжя', 'Мелітополь', 'Бердянськ', 'Енергодар'
            ],
            'Вінницька область': [
                'Вінниця', 'Жмеринка', 'Могилів-Подільський'
            ],
            'Чернігівська область': [
                'Чернігів', 'Ніжин', 'Прилуки'
            ],
            'Чернівецька область': [
                'Чернівці', 'Сторожинець'
            ],
            'Житомирська область': [
                'Житомир', 'Бердичів', 'Коростень'
            ],
            'Сумська область': [
                'Суми', 'Конотоп', 'Шостка'
            ],
            'Рівненська область': [
                'Рівне', 'Дубно', 'Костопіль'
            ],
            'Івано-Франківська область': [
                'Івано-Франківськ', 'Калуш', 'Коломия'
            ],
            'Херсонська область': [
                'Херсон', 'Нова Каховка', 'Каховка'
            ],
            'Тернопільська область': [
                'Тернопіль', 'Кременець', 'Чортків'
            ],
            'Хмельницька область': [
                'Хмельницький', 'Каменець-Подільський', 'Шепетівка'
            ],
            'Волинська область': [
                'Луцьк', 'Ковель', 'Нововолинськ'
            ],
            'Закарпатська область': [
                'Ужгород', 'Мукачево', 'Хуст'
            ],
            'Луганська область': [
                'Луганськ', 'Сєвєродонецьк', 'Алчевськ', 'Лисичанськ'
            ],
            'Донецька область': [
                'Донецьк', 'Маріуполь', 'Горлівка', 'Краматорськ'
            ],
        }

    def normalize_text(self, text: str) -> str:
        """Нормализация текста для сравнения (такая же как в API)"""
        if not isinstance(text, str):
            text = str(text)

        # Приводим к нижнему регистру
        text = text.lower()

        # Убираем лишние пробелы и неразрывные пробелы
        text = ' '.join(text.split())
        text = text.replace('\xa0', ' ').replace('\u200b', '').replace('\u200e', '').replace('\u200f', '')

        # Убираем двойные пробелы
        while '  ' in text:
            text = text.replace('  ', ' ')

        # Убираем лишние точки и запятые
        text = text.replace('..', '.').replace(',,', ',')

        # Оставляем только буквы, цифры, пробелы, скобки, тире, плюс, точку и запятую
        text = re.sub(r'[^\w\s()\-+.,]', '', text, flags=re.UNICODE)

        # Нормализуем скобки - убираем пробелы внутри скобок
        text = re.sub(r'\(\s+', '(', text)
        text = re.sub(r'\s+\)', ')', text)

        # Нормализуем пробелы вокруг знаков
        text = re.sub(r'\s*-\s*', '-', text)
        text = re.sub(r'\s*\+\s*', '+', text)
        text = re.sub(r'\s*,\s*', ',', text)
        text = re.sub(r'\s*\.\s*', '.', text)

        return text.strip()

    def load_work_types(self, file_path: str = None):
        """Загружает типы работ из файла или базы данных"""
        try:
            # Очищаем словарь
            self.work_types_dict = {}

            if file_path and os.path.exists(file_path):
                # Загружаем из файла
                self.work_types_df = pd.read_excel(file_path)
                print(f"Загружено {len(self.work_types_df)} типов работ из файла")

                # Сохраняем в базу данных
                self.save_work_types_to_db(self.work_types_df)

            # Всегда загружаем из базы данных (даже если был файл)
            print("Загрузка типов работ из базы данных...")

            with get_db_cursor() as cursor:
                cursor.execute("""
                    SELECT name, has_writeoff_materials, has_writeoff_equipment, 
                           demount_lines_count, is_to
                    FROM work_types
                """)

                rows = cursor.fetchall()
                print(f"Загружено {len(rows)} типов работ из БД")

                for row in rows:
                    name = str(row['name']).strip()
                    name = ' '.join(name.split())
                    name = name.replace('\xa0', ' ')

                    # Создаем несколько вариантов ключей для поиска
                    normalized_name = self.normalize_text(name)

                    # Основной ключ
                    self.work_types_dict[normalized_name] = {
                        'original_name': name,
                        'has_writeoff_materials': bool(row['has_writeoff_materials']),
                        'has_writeoff_equipment': bool(row['has_writeoff_equipment']),
                        'demount_lines_count': int(row['demount_lines_count']),
                        'is_to': bool(row['is_to'])
                    }

                    # Дополнительные ключи для поиска
                    # Убираем (АСТ) для поиска
                    name_without_ast = normalized_name.replace('(аст)', '').strip()
                    if name_without_ast and name_without_ast not in self.work_types_dict:
                        self.work_types_dict[name_without_ast] = {
                            'original_name': name,
                            'has_writeoff_materials': bool(row['has_writeoff_materials']),
                            'has_writeoff_equipment': bool(row['has_writeoff_equipment']),
                            'demount_lines_count': int(row['demount_lines_count']),
                            'is_to': bool(row['is_to'])
                        }

                # Выводим примеры для отладки
                print("\nПримеры загруженных типов работ (первые 5):")
                for i, (key, value) in enumerate(list(self.work_types_dict.items())[:5]):
                    print(f"  {i + 1}. Ключ: '{key}' -> Имя: '{value['original_name']}' (ТО: {value['is_to']})")

            print(f"\nВсего загружено типов работ: {len(self.work_types_dict)}")

        except Exception as e:
            print(f"Ошибка загрузки типов работ: {e}")
            import traceback
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=f"Ошибка загрузки типов работ: {e}")

    def save_work_types_to_db(self, work_types_df):
        """Сохраняет типы работ в базу данных"""
        try:
            with get_db_cursor() as cursor:
                # Очищаем таблицу
                cursor.execute("DELETE FROM work_types")

                for _, row in work_types_df.iterrows():
                    name = str(row.get('Наименование', '')).strip()
                    if not name:
                        continue

                    normalized_name = self.normalize_text(name)

                    # Конвертируем значения
                    def safe_int(value, default=0):
                        try:
                            if pd.isna(value):
                                return default
                            return int(float(value))
                        except:
                            return default

                    cursor.execute("""
                        INSERT INTO work_types 
                        (name, normalized_name, has_writeoff_materials, 
                         has_writeoff_equipment, demount_lines_count, is_to,
                         created_at, updated_at)
                        VALUES (?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                    """, (
                        name,
                        normalized_name,
                        safe_int(row.get('Наличие списания материалов', 0)),
                        safe_int(row.get('Наличие списания оборудования', 0)),
                        safe_int(row.get('Количество строк демонтажа', 0)),
                        safe_int(row.get('ТО', 0))
                    ))

                print(f"Сохранено {len(work_types_df)} типов работ в БД")

        except Exception as e:
            print(f"Ошибка сохранения типов работ в БД: {e}")
            raise

    def load_orders(self, file_path: str):
        """Загружает файл с заявками"""
        try:
            # Читаем файл
            self.orders_df = pd.read_excel(file_path)
            print(f"Загружен файл с {len(self.orders_df)} строками")
            print(f"Колонки в файле: {list(self.orders_df.columns)}")

            # Переименовываем колонки для унификации
            column_mapping = {
                'Тип работ': 'Тип работ',
                'Тип_работ': 'Тип работ',
                'Номер заявки': 'Номер заявки',
                'Номер_заявки': 'Номер заявки',
                'Город': 'Город',
                'Исполнитель': 'Исполнитель',
                'Наличие списания материалов': 'Наличие списания материалов',
                'Наличие_списания_материалов': 'Наличие списания материалов',
                'Наличие списания оборудования': 'Наличие списания оборудования',
                'Наличие_списания_оборудования': 'Наличие списания оборудования',
                'Количество строк демонтажа': 'Количество строк демонтажа',
                'Количество_строк_демонтажа': 'Количество строк демонтажа',
                'Наряд': 'Наряд'
            }

            # Применяем переименование
            for old_col, new_col in column_mapping.items():
                if old_col in self.orders_df.columns and new_col not in self.orders_df.columns:
                    self.orders_df.rename(columns={old_col: new_col}, inplace=True)
                    print(f"  Переименовано: '{old_col}' -> '{new_col}'")

            # Проверяем обязательные колонки
            required_columns = ['Тип работ', 'Номер заявки', 'Город']

            # Проверяем какие колонки есть
            existing_columns = []
            missing_columns = []
            for col in required_columns:
                if col in self.orders_df.columns:
                    existing_columns.append(col)
                else:
                    missing_columns.append(col)

            print(f"  Найдены колонки: {existing_columns}")

            if missing_columns:
                print(f"  Отсутствуют колонки: {missing_columns}")
                print(f"  Все доступные колонки: {list(self.orders_df.columns)}")
                raise HTTPException(
                    status_code=400,
                    detail=f"В файле заявок отсутствуют обязательные колонки: {', '.join(missing_columns)}"
                )

            # Конвертируем числовые колонки
            numeric_columns = ['Наличие списания материалов', 'Наличие списания оборудования',
                               'Количество строк демонтажа']
            for col in numeric_columns:
                if col in self.orders_df.columns:
                    # Заменяем NaN на 0
                    self.orders_df[col] = self.orders_df[col].fillna(0)
                    # Конвертируем в int
                    self.orders_df[col] = pd.to_numeric(self.orders_df[col], errors='coerce').fillna(0).astype(int)
                    print(f"  Обработана числовая колонка: '{col}'")

            # Проверяем первые несколько строк для отладки
            print(f"\nПримеры данных (первые 3 строки):")
            for i in range(min(3, len(self.orders_df))):
                row = self.orders_df.iloc[i]
                print(
                    f"  Строка {i}: Тип работ='{row.get('Тип работ', '')}', Город='{row.get('Город', '')}', Материалы={row.get('Наличие списания материалов', 0)}")

            print(f"\nЗагружено {len(self.orders_df)} заявок из файла")

        except Exception as e:
            print(f"Ошибка загрузки файла заявок: {e}")
            import traceback
            traceback.print_exc()
            raise HTTPException(status_code=400, detail=f"Ошибка загрузки файла заявок: {e}")

    def find_work_type(self, work_type_name: str) -> Optional[Dict]:
        """Найти тип работ по названию"""
        if not work_type_name or pd.isna(work_type_name):
            return None

        work_type_str = str(work_type_name).strip()
        if not work_type_str:
            return None

        # Нормализуем название
        normalized_search = self.normalize_text(work_type_str)
        print(f"Поиск типа: '{work_type_str}' -> нормализовано: '{normalized_search}'")

        # 1. Прямое совпадение
        if normalized_search in self.work_types_dict:
            print(f"  ✓ Прямое совпадение")
            return self.work_types_dict[normalized_search]

        # 2. Пробуем убрать (аст)
        if '(аст)' in normalized_search:
            search_without_ast = normalized_search.replace('(аст)', '').strip()
            if search_without_ast in self.work_types_dict:
                print(f"  ✓ Совпадение без (АСТ)")
                return self.work_types_dict[search_without_ast]

        # 3. Ищем частичное совпадение
        for key, work_type in self.work_types_dict.items():
            # Убираем (аст) из обоих строк для сравнения
            search_clean = normalized_search.replace('(аст)', '').strip()
            key_clean = key.replace('(аст)', '').strip()

            if search_clean == key_clean:
                print(f"  ✓ Совпадение после очистки (АСТ)")
                return work_type

            # Частичное вхождение
            if search_clean in key_clean or key_clean in search_clean:
                # Вычисляем степень совпадения
                shorter = min(len(search_clean), len(key_clean))
                longer = max(len(search_clean), len(key_clean))

                if shorter > 0:
                    similarity = shorter / longer
                    if similarity > 0.8:  # 80% совпадение
                        print(f"  ✓ Частичное совпадение ({similarity:.0%})")
                        return work_type

        print(f"  ✗ Не найдено")
        return None

    def check_violations(self):
        """Проверяет заявки на нарушения правил"""
        # Проверяем, загружены ли данные
        if not self.work_types_dict or self.orders_df is None or len(self.orders_df) == 0:
            raise HTTPException(
                status_code=400,
                detail="Сначала загрузите типы работ и заявки. work_types_dict: {}, orders_df: {}".format(
                    len(self.work_types_dict) if self.work_types_dict else 0,
                    len(self.orders_df) if self.orders_df is not None else 0
                )
            )

        # Сбрасываем списки
        self.violation_orders = []
        self.to_violations = []
        self.non_to_violations = []

        # Статистика
        matched_types = set()
        unmatched_types = set()
        skipped_orders = 0

        print(f"\nНачало проверки {len(self.orders_df)} заявок...")

        for idx, row in self.orders_df.iterrows():
            try:
                # Получаем тип работ
                work_type_cell = row.get('Тип работ')
                if pd.isna(work_type_cell):
                    skipped_orders += 1
                    continue

                work_type_name = str(work_type_cell).strip()

                if not work_type_name or work_type_name.lower() == 'nan':
                    skipped_orders += 1
                    continue

                # Ищем тип работ
                work_type = self.find_work_type(work_type_name)

                if work_type:
                    matched_types.add(work_type_name)

                    # Получаем фактические значения
                    actual_materials = 0
                    actual_equipment = 0
                    actual_demount = 0

                    # Безопасное получение значений материалов
                    materials_cell = row.get('Наличие списания материалов')
                    if pd.notna(materials_cell):
                        try:
                            actual_materials = int(float(materials_cell))
                        except (ValueError, TypeError):
                            actual_materials = 0

                    # Безопасное получение значений оборудования
                    equipment_cell = row.get('Наличие списания оборудования')
                    if pd.notna(equipment_cell):
                        try:
                            actual_equipment = int(float(equipment_cell))
                        except (ValueError, TypeError):
                            actual_equipment = 0

                    # Безопасное получение значений демонтажа
                    demount_cell = row.get('Количество строк демонтажа')
                    if pd.notna(demount_cell):
                        try:
                            actual_demount = int(float(demount_cell))
                        except (ValueError, TypeError):
                            actual_demount = 0

                    # Проверяем нарушения
                    is_violation = False
                    violations = []

                    print(f"\nПроверка заявки {idx}:")
                    print(f"  Тип: '{work_type_name}' -> '{work_type['original_name']}' (ТО: {work_type['is_to']})")
                    print(
                        f"  Правила: материалы={work_type['has_writeoff_materials']}, оборудование={work_type['has_writeoff_equipment']}, демонтаж={work_type['demount_lines_count']}")
                    print(
                        f"  Факт: материалы={actual_materials}, оборудование={actual_equipment}, демонтаж={actual_demount}")

                    # Общие правила для ВСЕХ заявок (ТО и не-ТО)

                    # 1. Проверка материалов
                    if work_type['has_writeoff_materials'] == 1 and actual_materials == 0:
                        is_violation = True
                        violations.append("Нет списания материалов")
                        print(f"  ✓ Нет списания материалов -> НАРУШЕНИЕ")

                    # 2. Проверка оборудования
                    if work_type['has_writeoff_equipment'] == 1 and actual_equipment == 0:
                        is_violation = True
                        violations.append("Нет списания оборудования")
                        print(f"  ✓ Нет списания оборудования -> НАРУШЕНИЕ")

                    # 3. Проверка строк демонтажа
                    if work_type['demount_lines_count'] == 1 and actual_demount == 0:
                        is_violation = True
                        violations.append("Нет строк демонтажа")
                        print(f"  ✓ Нет строк демонтажа -> НАРУШЕНИЕ")

                    # Дополнительное правило для ТО заявок:
                    # Если тип ТО, но в правилах НЕТ ни материалов, ни оборудования - это нарушение
                    if work_type['is_to']:
                        if work_type['has_writeoff_materials'] == 0 and work_type['has_writeoff_equipment'] == 0:
                            is_violation = True
                            violations.append("ТО заявка без списания материалов и оборудования")
                            print(f"  ✓ ТО заявка без списания материалов и оборудования -> НАРУШЕНИЕ")

                    if not violations:
                        print("  ✓ Нет нарушений")

                    if is_violation:
                        violation_data = {
                            'index': idx,
                            'Наряд': str(row.get('Наряд', '')) if 'Наряд' in row and pd.notna(row.get('Наряд')) else '',
                            'Исполнитель': str(row.get('Исполнитель', '')) if 'Исполнитель' in row and pd.notna(
                                row.get('Исполнитель')) else '',
                            'Город': str(row.get('Город', '')) if 'Город' in row and pd.notna(row.get('Город')) else '',
                            'Номер заявки': str(row.get('Номер заявки', '')) if 'Номер заявки' in row and pd.notna(
                                row.get('Номер заявки')) else '',
                            'Тип работ': work_type_name,
                            'Тип работ_оригинал_из_правил': work_type['original_name'],
                            'is_to': work_type['is_to'],
                            'rules': {
                                'has_writeoff_materials': work_type['has_writeoff_materials'],
                                'has_writeoff_equipment': work_type['has_writeoff_equipment'],
                                'demount_lines_count': work_type['demount_lines_count']
                            },
                            'actual': {
                                'materials': actual_materials,
                                'equipment': actual_equipment,
                                'demount': actual_demount
                            },
                            'violations': violations
                        }

                        self.violation_orders.append(violation_data)

                        if work_type['is_to']:
                            self.to_violations.append(violation_data)
                        else:
                            self.non_to_violations.append(violation_data)

                else:
                    unmatched_types.add(work_type_name)
                    print(f"  ✗ Тип не найден: '{work_type_name}'")

            except Exception as e:
                print(f"Ошибка проверки строки {idx}: {e}")
                import traceback
                traceback.print_exc()
                skipped_orders += 1
                continue

        # Формируем статистику
        stats = {
            'total_checked': len(self.orders_df),
            'matched_types': len(matched_types),
            'unmatched_types': len(unmatched_types),
            'skipped_orders': skipped_orders,
            'violation_orders': len(self.violation_orders),
            'to_violations': len(self.to_violations),
            'non_to_violations': len(self.non_to_violations),
            'matched_types_list': sorted(list(matched_types)),
            'unmatched_types_list': sorted(list(unmatched_types))
        }

        print("\n" + "=" * 60)
        print("Статистика проверки:")
        print("=" * 60)
        print(f"  Всего проверено заявок: {stats['total_checked']}")
        print(f"  Сопоставлено типов работ: {stats['matched_types']}")
        print(f"  Не сопоставлено типов работ: {stats['unmatched_types']}")
        print(f"  Пропущено заявок: {stats['skipped_orders']}")
        print(f"  Всего нарушений: {stats['violation_orders']}")
        print(f"  ТО заявки с нарушениями: {stats['to_violations']}")
        print(f"  Не-ТО заявки с нарушениями: {stats['non_to_violations']}")

        if stats['matched_types_list']:
            print(f"\nСопоставленные типы ({len(stats['matched_types_list'])}):")
            for i, wt in enumerate(stats['matched_types_list'][:10], 1):
                print(f"  {i}. {wt}")
            if len(stats['matched_types_list']) > 10:
                print(f"  ... и еще {len(stats['matched_types_list']) - 10}")

        if stats['unmatched_types_list']:
            print(f"\nНе сопоставленные типы ({len(stats['unmatched_types_list'])}):")
            for i, wt in enumerate(stats['unmatched_types_list'][:10], 1):
                print(f"  {i}. {wt}")
            if len(stats['unmatched_types_list']) > 10:
                print(f"  ... и еще {len(stats['unmatched_types_list']) - 10}")

        return stats


    def highlight_orders_in_file(self, original_file_path: str, output_dir: str) -> Dict:
        """Закрашивает заявки в файле: ТО - красным, не-ТО - желтым"""
        try:
            wb = openpyxl.load_workbook(original_file_path)
            ws = wb.active

            # Находим строку с заголовками
            header_row = 0
            for row in range(1, 20):
                for col in range(1, ws.max_column + 1):
                    cell_value = str(ws.cell(row=row, column=col).value)
                    if 'Тип работ' in cell_value or 'Наряд' in cell_value:
                        header_row = row
                        break
                if header_row > 0:
                    break

            if header_row == 0:
                header_row = 1

            # Создаем заливки
            red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
            yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')

            # Создаем быстрый доступ к нарушениям
            violations_by_index = {v['index']: v for v in self.violation_orders}

            to_count = 0
            non_to_count = 0

            # Закрашиваем строки
            for excel_row in range(header_row + 1, ws.max_row + 1):
                pandas_idx = excel_row - header_row - 1

                if pandas_idx in violations_by_index:
                    violation = violations_by_index[pandas_idx]

                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row=excel_row, column=col)

                        if violation['is_to']:
                            cell.fill = red_fill
                            if col == 1:
                                to_count += 1
                        else:
                            cell.fill = yellow_fill
                            if col == 1:
                                non_to_count += 1

            # Сохраняем файл
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = os.path.join(output_dir, f'highlighted_orders_{timestamp}.xlsx')
            wb.save(output_file)

            highlight_stats = {
                'output_file': output_file,
                'to_highlighted': to_count,
                'non_to_highlighted': non_to_count
            }

            return highlight_stats

        except Exception as e:
            print(f"Ошибка при выделении строк: {e}")
            raise

    def save_to_violations_file(self, output_dir: str) -> str:
        """Сохраняет ТО заявки с нарушениями в отдельный файл"""
        if not self.to_violations:
            return None

        try:
            to_data = []
            for violation in self.to_violations:
                to_data.append({
                    'Номер заявки': violation['Номер заявки'],
                    'Город': violation['Город'],
                    'Исполнитель': violation['Исполнитель'],
                    'Тип работ (в заявке)': violation['Тип работ'],
                    'Тип работ (в правилах)': violation['Тип работ_оригинал_из_правил'],
                    'Наличие списания материалов (факт)': violation['actual']['materials'],
                    'Наличие списания оборудования (факт)': violation['actual']['equipment'],
                    'Нарушения': ', '.join(violation['violations'])
                })

            to_df = pd.DataFrame(to_data)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = os.path.join(output_dir, f'to_violations_{timestamp}.xlsx')
            to_df.to_excel(output_file, index=False)

            return output_file

        except Exception as e:
            print(f"Ошибка сохранения ТО заявок: {e}")
            return None

    def group_non_to_violations_by_region(self) -> Dict[str, List]:
        """Группирует не-ТО заявки с нарушениями по областям"""
        if not self.non_to_violations:
            return {}

        grouped_by_region = defaultdict(list)

        for violation in self.non_to_violations:
            city = str(violation['Город']).strip() if violation['Город'] else ''
            region = self._find_region(city)

            grouped_by_region[region].append({
                'Номер заявки': violation['Номер заявки'],
                'Город': city,
                'Исполнитель': violation['Исполнитель'],
                'Тип работ': violation['Тип работ']
            })

        return grouped_by_region

    def _find_region(self, city: str) -> str:
        """Определяет область по городу"""
        if not city or city.lower() == 'nan' or city == '':
            return 'Не указан'

        city_str = city.lower().strip()

        # Специальная проверка для Кременчука и Горішніх Плавнів
        if 'кременчук' in city_str:
            return 'Кременчук і Горішні Плавні (окремо)'
        if 'горішні' in city_str and 'плавні' in city_str:
            return 'Кременчук і Горішні Плавні (окремо)'

        # Проверяем все области
        for region, cities in self.regions.items():
            for region_city in cities:
                if region_city.lower() in city_str:
                    return region

        return 'Інші регіони'

    def save_non_to_to_file(self, output_dir: str, grouped_data: Dict[str, List]) -> str:
        """Сохраняет сгруппированные не-ТО заявки в текстовый файл"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = os.path.join(output_dir, f'non_to_violations_{timestamp}.txt')

            with open(output_file, 'w', encoding='utf-8') as f:
                # Порядок вывода областей
                priority_order = [
                    'Кременчук і Горішні Плавні (окремо)',
                    'Полтавська область',
                    'Одеська область',
                    'Миколаївська область',
                    'Черкаська область',
                    'Кіровоградська область'
                ]

                other_regions = [r for r in grouped_data.keys() if r not in priority_order]
                sorted_regions = priority_order + sorted(other_regions)

                for region in sorted_regions:
                    if region not in grouped_data or not grouped_data[region]:
                        continue

                    orders = grouped_data[region]
                    f.write(f"{region.upper()}:\n")
                    f.write("Привет, не списаны материалы за\n\n")

                    sorted_orders = sorted(orders, key=lambda x: x['Номер заявки'])

                    for order in sorted_orders:
                        f.write(f"{order['Номер заявки']} - {order['Город']} - {order['Исполнитель']}\n")

                    f.write(f"\nВсего: {len(orders)} заявок\n\n")
                    f.write("=" * 60 + "\n\n")

            return output_file

        except Exception as e:
            print(f"Ошибка сохранения файла: {e}")
            return None

    def save_non_to_regional_files(self, output_dir: str, grouped_data: Dict[str, List]) -> Dict[str, str]:
        """Сохраняет не-ТО заявки в отдельные файлы по областям"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            files = {}

            priority_order = [
                'Кременчук і Горішні Плавні (окремо)',
                'Полтавська область',
                'Одеська область',
                'Миколаївська область',
                'Черкаська область',
                'Кіровоградська область'
            ]

            other_regions = [r for r in grouped_data.keys() if r not in priority_order]
            sorted_regions = priority_order + sorted(other_regions)

            for region in sorted_regions:
                if region not in grouped_data or not grouped_data[region]:
                    continue

                safe_region_name = re.sub(r'[^\w\-_\. ]', '_', region)
                region_file = os.path.join(output_dir, f'{safe_region_name}_{timestamp}.txt')

                orders = grouped_data[region]

                with open(region_file, 'w', encoding='utf-8') as f:
                    f.write(f"{region.upper()}:\n")
                    f.write("Привет, не списаны материалы за\n\n")

                    sorted_orders = sorted(orders, key=lambda x: x['Номер заявки'])

                    for order in sorted_orders:
                        f.write(f"{order['Номер заявки']} - {order['Город']} - {order['Исполнитель']}\n")

                    f.write(f"\nВсего: {len(orders)} заявок\n\n")
                    f.write("=" * 60 + "\n\n")

                files[region] = region_file

            return files

        except Exception as e:
            print(f"Ошибка сохранения региональных файлов: {e}")
            return {}

    def create_summary_report(self, stats: Dict, output_dir: str) -> str:
        """Создает сводный отчет по проверке"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            report_file = os.path.join(output_dir, f'summary_report_{timestamp}.txt')

            with open(report_file, 'w', encoding='utf-8') as f:
                f.write("=" * 60 + "\n")
                f.write("СВОДНЫЙ ОТЧЕТ ПО ПРОВЕРКЕ ЗАЯВОК\n")
                f.write("=" * 60 + "\n\n")

                f.write(f"Дата проверки: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}\n\n")

                f.write("СТАТИСТИКА:\n")
                f.write("-" * 40 + "\n")
                f.write(f"Всего проверено заявок: {stats['total_checked']}\n")
                f.write(f"Сопоставлено типов работ: {stats['matched_types']}\n")
                f.write(f"Не сопоставлено типов работ: {stats['unmatched_types']}\n")
                f.write(f"Заявок с нарушениями: {stats['violation_orders']}\n")
                f.write(f"  • ТО заявок с нарушениями: {stats['to_violations']}\n")
                f.write(f"  • Не-ТО заявок с нарушениями: {stats['non_to_violations']}\n\n")

                if stats['matched_types_list']:
                    f.write("СОПОСТАВЛЕННЫЕ ТИПЫ РАБОТ:\n")
                    f.write("-" * 40 + "\n")
                    for i, work_type in enumerate(stats['matched_types_list'], 1):
                        f.write(f"{i}. {work_type}\n")

                if stats['unmatched_types_list']:
                    f.write("\nНЕ СОПОСТАВЛЕННЫЕ ТИПЫ РАБОТ:\n")
                    f.write("-" * 40 + "\n")
                    for i, work_type in enumerate(stats['unmatched_types_list'], 1):
                        f.write(f"{i}. {work_type}\n")

                f.write("\n" + "=" * 60 + "\n")
                f.write("ПРОВЕРКА ЗАВЕРШЕНА\n")
                f.write("=" * 60 + "\n")

            return report_file

        except Exception as e:
            print(f"Ошибка создания отчета: {e}")
            return None